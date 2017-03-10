# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import os
import pandas
import datetime
import xlsxwriter
import openpyxl
import xlrd


os.chdir(r'\Users\Ben\Documents\GitHub\Ben-Projects')

def make_excel_file(title):
        workbook = xlsxwriter.Workbook(title)
        worksheet = workbook.add_worksheet()
        worksheet.write(0,0, "JOURNAL")
        worksheet.write(0,1, "DATE")
        worksheet.write(0,2, "ACCOUNT")
        worksheet.write(0,3, "CONTROL #")
        worksheet.write(0,4, "DOCUMENT #")
        worksheet.write(0,5, "REFERENCE #")
        worksheet.write(0,6, "AMOUNT")
        worksheet.write(0,7, "DESCRIPTION")
        return workbook.filename
        
def add_row(wb1,row, column, date,account,controlnumber,documentnumber,amount,description):
       book = openpyxl.load_workbook(wb1)
       sheet = book.active
       sheet.cell(row = row,column = column).value = "GJE"
       sheet.cell(row = row,column = column +1).value  = date
       sheet.cell(row = row,column = column +2).value  = account
       sheet.cell(row = row,column = column +3).value = controlnumber
       sheet.cell(row = row,column = column +4).value  = documentnumber
       sheet.cell(row = row,column = column +5).value  = documentnumber
       sheet.cell(row = row,column = column +6).value = amount
       sheet.cell(row = row,column = column +7).value  = description
       book.save(wb1)

wb1 = make_excel_file('testing.xlsx')
add_row(wb1,2,1,'2/21/17','300','SUB001','TESTING','5000','THIS IS A TEST')

